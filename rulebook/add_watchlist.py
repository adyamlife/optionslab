from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F4E78")
NOTE_FONT = Font(name=FONT, italic=True, color="666666")
BODY_FONT = Font(name=FONT)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

path = r"C:\Project Y\rulebook\Options_Trading_Rulebook.xlsx"
wb = load_workbook(path)

# Insert Watchlist sheet right after "8. Open Items" (or at end)
ws = wb.create_sheet("0. Watchlist")
wb.move_sheet(ws, offset=-(len(wb.sheetnames) - 1))  # move to front

ws.cell(row=1, column=1, value="WATCHLIST (handful of underlyings only - not the full market)").font = TITLE_FONT
ws.cell(row=2, column=1, value="Keep this list short. Add/remove based on quarterly review (liquidity, IV behavior, your familiarity with the name).").font = NOTE_FONT

headers = ["Underlying", "Asset Type", "Typical IV Rank Range",
           "Liquidity OK? (Y/N)", "Avg Bid-Ask Spread %", "Sector/Notes",
           "Earnings Date (next)", "Active? (Y/N)"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=4, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

for r in range(5, 15):
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = BORDER

widths = [14, 14, 20, 16, 18, 35, 16, 12]
for c, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A5"

wb.save(path)
print("saved")
