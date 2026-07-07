from openpyxl import load_workbook

path = r"C:\Project Y\rulebook\Options_Trading_Rulebook.xlsx"
wb = load_workbook(path)
ws = wb["0. Watchlist"]

# row -> (Typical IV Rank Range, Liquidity OK, Avg Bid-Ask Spread %, Sector/Notes, Earnings Date)
data = {
    5: ("N/A (ETF)", "Y", "Very tight (~$0.01-0.02)", "S&P 500 index ETF - benchmark liquidity, use as the baseline", "N/A (ETF, no earnings)"),
    6: ("N/A (ETF)", "Y", "Tight, slightly wider than QQQ/SPY", "3x leveraged Nasdaq ETF - high IV due to leverage/decay; verify spread width fits $120 max-loss rule before using", "N/A (ETF, no earnings)"),
    7: ("N/A (ETF)", "Y", "Very tight (~$0.01-0.02)", "Nasdaq-100 index ETF - high liquidity, good for credit spreads", "N/A (ETF, no earnings)"),
    8: ("Check broker daily - changes with market conditions", "Y", "Tight (mega-cap, very liquid chain)", "Mega-cap tech - reliable liquidity but premiums often lower (low IV) outside earnings", "Jul 30, 2026 (confirmed, after close)"),
    9: ("Check broker daily - changes with market conditions", "Y (verify)", "Moderate - wider than mega-caps, check before entry", "Mid-cap, more volatile than AAPL - good IV but confirm OI/spread on chosen strikes", "Jul 29, 2026 (estimate, before open - confirm closer to date)"),
    10: ("Check broker daily - changes with market conditions", "Y (verify)", "Moderate - check before entry", "Mid-cap retail - earnings-sensitive, confirm liquidity on monthly vs weekly expiries", "Aug 27, 2026 (estimate)"),
}

for row, (iv, liq, spread, notes, earn) in data.items():
    ws.cell(row=row, column=3, value=iv)
    ws.cell(row=row, column=4, value=liq)
    ws.cell(row=row, column=5, value=spread)
    ws.cell(row=row, column=6, value=notes)
    ws.cell(row=row, column=7, value=earn)

wb.save(path)
print("saved")
