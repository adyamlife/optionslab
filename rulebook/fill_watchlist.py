from openpyxl import load_workbook

path = r"C:\Project Y\rulebook\Options_Trading_Rulebook.xlsx"
wb = load_workbook(path)
ws = wb["0. Watchlist"]

entries = [
    ("SPY", "ETF (Index)"),
    ("TQQQ", "ETF (3x Leveraged)"),
    ("QQQ", "ETF (Index)"),
    ("AAPL", "Stock"),
    ("SPOT", "Stock"),
    ("ULTA", "Stock"),
]

for i, (ticker, atype) in enumerate(entries):
    row = 5 + i
    ws.cell(row=row, column=1, value=ticker)
    ws.cell(row=row, column=2, value=atype)
    ws.cell(row=row, column=8, value="Y")

wb.save(path)
print("saved")
