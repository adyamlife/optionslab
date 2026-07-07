from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F4E78")
NOTE_FONT = Font(name=FONT, italic=True, color="666666")
BODY_FONT = Font(name=FONT)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def write_table(ws, start_row, headers, rows, col_widths, title=None, note=None):
    r = start_row
    if title:
        ws.cell(row=r, column=1, value=title).font = TITLE_FONT
        r += 1
    if note:
        ws.cell(row=r, column=1, value=note).font = NOTE_FONT
        r += 1
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    r += 1
    for row in rows:
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = BORDER
        r += 1
    for c, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    return r + 1

# ---------------- Sheet 1: Universal Rules ----------------
ws = wb.active
ws.title = "1. Universal Rules"
ws.cell(row=1, column=1, value="UNIVERSAL RULES (apply to every trade, regardless of structure)").font = TITLE_FONT
rows = [
    ["Max risk per trade", "Max loss <= 12% of capital (~$120 on $1000 account), known BEFORE entry", "Limits damage from any single wrong trade"],
    ["Max concurrent positions", "2 open positions max (given $1000 capital)", "Avoids over-extension / correlated blowups"],
    ["Event blackout", "No new entry if earnings or major macro event (Fed/CPI/jobs) falls within the planned holding period", "Avoids predictable, avoidable losses"],
    ["Liquidity gate", "Only trade strikes where bid-ask spread <= 5-10% of mid price and OI is meaningful", "Avoids slippage eating the edge"],
    ["Holding period", "Target 5-10 DTE at entry. Avoid <3 DTE (gamma risk) and >21 DTE (capital tied up too long)", "Balances theta decay vs gamma/gap risk"],
    ["No-trade is valid", "If situation doesn't clearly match a structure in the matrix, skip the week for that underlying", "Forcing trades is a leading cause of losses"],
]
write_table(ws, 3, ["Rule", "Definition", "Why"], rows, [22, 60, 45])

# ---------------- Sheet 2: Situation Assessment ----------------
ws = wb.create_sheet("2. Situation Assessment")
ws.cell(row=1, column=1, value="SITUATION ASSESSMENT (do this first, every time)").font = TITLE_FONT
rows = [
    ["IV environment", "IV rank / percentile of underlying vs its own 1-yr range", "High (>50th pct) / Low (<50th pct)"],
    ["Trend", "Price vs 20-day and 50-day moving average, direction of both", "Uptrend / Downtrend / Range-bound"],
    ["Event proximity", "Earnings / macro calendar within the planned holding period", "Clear / Blocked"],
]
write_table(ws, 3, ["Axis", "How to check", "Categories"], rows, [20, 55, 35],
             note="If 'Blocked' on Event proximity -> no trade this underlying this week, regardless of IV/trend.")

# ---------------- Sheet 3: Structure Selection Matrix ----------------
ws = wb.create_sheet("3. Structure Matrix")
ws.cell(row=1, column=1, value="STRUCTURE SELECTION MATRIX").font = TITLE_FONT
rows = [
    ["High", "Range-bound", "Iron Condor (sell both sides, defined risk)", "Premium is rich, no directional edge needed"],
    ["High", "Uptrend", "Put credit spread (sell put side only)", "Sell rich premium with the trend for extra cushion"],
    ["High", "Downtrend", "Call credit spread (sell call side only)", "Same logic, opposite direction"],
    ["Low", "Uptrend", "Call debit spread (buy near-the-money call, sell further OTM call)", "Premium is cheap - buy it instead of selling it"],
    ["Low", "Downtrend", "Put debit spread (buy near-the-money put, sell further OTM put)", "Premium is cheap - buy it instead of selling it"],
    ["Low", "Range-bound", "NO TRADE (or calendar/diagonal - advanced, deferred)", "Cheap premium + no movement = poor risk/reward either way"],
]
write_table(ws, 3, ["IV Environment", "Trend", "Structure", "Rationale"], rows, [16, 14, 50, 50])

# ---------------- Sheet 4: Entry Criteria ----------------
ws = wb.create_sheet("4. Entry Criteria")
ws.cell(row=1, column=1, value="ENTRY CRITERIA (per structure)").font = TITLE_FONT
rows = [
    ["Credit spread / Iron Condor", "Short strike delta", "0.15 - 0.25 (far enough OTM for cushion, not negligible premium)"],
    ["Credit spread / Iron Condor", "Spread width", "Sized so max loss = position-size limit (~$120)"],
    ["Credit spread / Iron Condor", "Credit received", ">= ~25-33% of spread width; if less, skip the trade"],
    ["Debit spread", "Long strike delta", "0.40 - 0.55 (near-the-money for directional conviction)"],
    ["Debit spread", "Short (hedge) strike delta", "0.15 - 0.20 (far enough out to meaningfully reduce cost)"],
    ["Debit spread", "Max loss", "= premium paid = position-size limit"],
]
write_table(ws, 3, ["Structure", "Criterion", "Rule"], rows, [26, 26, 65])

# ---------------- Sheet 5: Exit / Adjustment Rules ----------------
ws = wb.create_sheet("5. Exit-Adjustment Rules")
ws.cell(row=1, column=1, value="EXIT / ADJUSTMENT RULES (per structure)").font = TITLE_FONT
rows = [
    ["Credit spread / Iron Condor", "Profit target", "Close at 50-65% of max profit (credit received)"],
    ["Credit spread / Iron Condor", "Defend trigger", "Loss reaches ~1.5x credit received OR short strike delta exceeds ~0.40 -> roll out/down for credit, or convert to condor/butterfly"],
    ["Credit spread / Iron Condor", "Time-based exit", "Close or roll by 2-3 DTE regardless of P&L"],
    ["Debit spread", "Profit target", "Close at 50-75% of max profit (spread width minus cost)"],
    ["Debit spread", "Loss trigger", "Close if position value drops to ~50% of cost paid"],
    ["Debit spread", "Time-based exit", "If underlying hasn't moved as expected by ~50% of time to expiry, consider closing (theta works against you)"],
]
write_table(ws, 3, ["Structure", "Trigger", "Action"], rows, [26, 22, 70])

# ---------------- Sheet 6: Pre-Trade Checklist ----------------
ws = wb.create_sheet("6. Pre-Trade Checklist")
ws.cell(row=1, column=1, value="PRE-TRADE CHECKLIST (run in order, every time)").font = TITLE_FONT
rows = [
    [1, "Calendar check", "Any blocking events (earnings/macro) in holding window?", ""],
    [2, "IV rank check", "Classify High / Low", ""],
    [3, "Trend check", "Classify Up / Down / Range-bound", ""],
    [4, "Structure lookup", "Match IV + Trend to Structure Matrix (sheet 3)", ""],
    [5, "Liquidity check", "Bid-ask spread and OI acceptable on relevant strikes?", ""],
    [6, "Position sizing", "Calculate size so max loss <= $120", ""],
    [7, "Enter trade", "Place order per Entry Criteria (sheet 4)", ""],
    [8, "Set alerts", "Set delta / P&L adjustment triggers per sheet 5", ""],
    [9, "Log trade", "Record in Trade Journal (sheet 7)", ""],
]
write_table(ws, 3, ["#", "Step", "Detail", "Done? (Y/N)"], rows, [5, 22, 65, 14])

# ---------------- Sheet 7: Trade Journal ----------------
ws = wb.create_sheet("7. Trade Journal")
ws.cell(row=1, column=1, value="TRADE JOURNAL").font = TITLE_FONT
ws.cell(row=2, column=1, value="Log every trade here. P&L and % Return are calculated automatically.").font = NOTE_FONT

headers = ["Date Entered", "Underlying", "IV Env", "Trend", "Structure", "Strikes",
           "Expiry", "Credit/Debit ($)", "Max Loss ($)", "Date Closed/Adjusted",
           "Exit Value ($)", "P&L ($)", "% Return on Risk", "Adjustment Made?", "Notes / Lessons"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=4, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

for r in range(5, 55):
    pl_cell = ws.cell(row=r, column=12)
    pl_cell.value = '=IF(J{0}="","",K{0}-H{0})'.format(r)
    pl_cell.font = BODY_FONT
    pl_cell.border = BORDER

    pct_cell = ws.cell(row=r, column=13)
    pct_cell.value = '=IF(OR(J{0}="",I{0}=0,I{0}=""),"",L{0}/I{0})'.format(r)
    pct_cell.font = BODY_FONT
    pct_cell.border = BORDER
    pct_cell.number_format = "0.0%"

    for c in range(1, 12):
        ws.cell(row=r, column=c).border = BORDER
    for c in (14, 15):
        ws.cell(row=r, column=c).border = BORDER

widths = [12, 12, 8, 10, 22, 16, 10, 14, 12, 16, 12, 10, 14, 14, 35]
for c, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.freeze_panes = "A5"

# ---------------- Sheet 8: Open Questions / Notes ----------------
ws = wb.create_sheet("8. Open Items")
ws.cell(row=1, column=1, value="OPEN QUESTIONS / NOTES TO REFINE").font = TITLE_FONT
rows = [
    ["Watchlist underlyings", "Which 1-2 stocks/ETFs to start with - affects realistic IV rank ranges, strike spacing, liquidity"],
    ["Data source for IV rank / delta", "Confirm broker platform provides this directly, or build a data-pulling script"],
    ["Liquidity thresholds", "Refine exact bid-ask spread % and OI minimums once looking at real chains"],
    ["Backtest validation", "Run rulebook against 1-2 years of historical chain data before committing real capital"],
]
write_table(ws, 3, ["Item", "Notes"], rows, [28, 90])

wb.save(r"C:\Project Y\rulebook\Options_Trading_Rulebook.xlsx")
print("saved")
